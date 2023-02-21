import sys
import subprocess
import requests
from bs4 import BeautifulSoup
import pandas as pd
import numpy as np
import matplotlib
from sklearn.linear_model import LinearRegression
from sklearn.metrics import r2_score
from sklearn.ensemble import RandomForestRegressor
from sklearn.model_selection import train_test_split
from sklearn.metrics import mean_absolute_error
from sklearn.feature_extraction.text import CountVectorizer
from sklearn.decomposition import LatentDirichletAllocation
import matplotlib.pyplot as plt
import seaborn as sns
import openpyxl
from openpyxl.chart import LineChart, Reference, Series
from selenium import webdriver
import anticaptcha


# Vérification des dépendances requises
required_packages = ['pandas', 'beautifulsoup4', 'selenium', 'openpyxl', 'scikit-learn','seaborn','sklearn']

for package in required_packages:
    try:
        __import__(package)
    except ImportError:
        subprocess.check_call([sys.executable, "-m", "pip", "install", package])
    else:
        subprocess.check_call([sys.executable, "-m", "pip", "install", "--upgrade", package])

# Fonction pour récupérer les informations d'un produit sur Interencheres
def get_product_info_interencheres(url, anti_captcha_key):
    # Récupération de la page web
    driver = webdriver.Chrome()
    driver.get(url)

    # Résolution du captcha AntiCaptcha
    if anti_captcha_key:
        token = anticaptcha.solve_captcha(anti_captcha_key, driver.current_url)
        driver.execute_script(f'document.getElementById("g-recaptcha-response").innerHTML="{token}";')
        driver.execute_script('document.getElementById("form_submit").submit();')

    # Récupération des informations du produit
    soup = BeautifulSoup(driver.page_source, 'html.parser')
    product_name = soup.find('div', {'class': 'vente-title'}).text
    product_price = soup.find('span', {'class': 'vente-prix'}).text
    product_description = soup.find('div', {'class': 'vente-description'}).text

    # Fermeture du navigateur
    driver.quit()

    # Stockage des informations dans un dictionnaire
    product = {
        'name': product_name,
        'price': product_price,
        'description': product_description
    }

    return product

# Fonction pour récupérer les informations d'un produit sur MoniteurLive
def get_product_info_moniteurlive(url):
    # Récupération de la page web
    page = requests.get(url)
    soup = BeautifulSoup(page.content, 'html.parser')

    # Récupération du titre, du prix de vente, de l'état du produit et des caractéristiques techniques
    product_title = soup.find('h1', {'class': 'ProductTitle'}).text
    product_price = soup.find('span', {'class': 'ProductPrice'}).text
    product_state = soup.find('div', {'class': 'ProductState'}).text
    tech_data = {}
    for item in soup.find_all('div', {'class': 'TechnicalData-item'}):
        key = item.find('div', {'class': 'TechnicalData-key'}).text
        value = item.find('div', {'class': 'TechnicalData-value'}).text
        tech_data[key] = value

    # Récupération du nom et de la note du vendeur
    seller_name = soup.find('div', {'class': 'SellerName'}).text
    seller_rating = soup.find('div', {'class': 'SellerRating'}).text

    # Récupération de la date de fin de l'enchère
    end_date = soup.find('div', {'class': 'AuctionEnd'}).find('span').text

    # Stockage des informations dans un dictionnaire
    product_info = {
        'title': product_title,
        'price': product_price,
        'state': product_state,
        'tech_data': tech_data,
        'seller_name': seller_name,
        'seller_rating': seller_rating,
        'end_date': end_date
    }

    return product_info


# Récupération des informations sur les produits
url_enchere = 'https://www.interencheres.com/'
anti_captcha_key = '9eecb67e989ef957d323678f518431eb'  # Remplacer par une clé AntiCaptcha valide
product_info_enchere = get_product_info_interencheres(url_enchere, anti_captcha_key)

url_moniteur = 'https://www.moniteurlive.com/auctions/future'
product_info_moniteur = get_product_info_moniteurlive(url_moniteur)

# Préparation des données pour l'entraînement du modèle
df = prepare_data([product_info_enchere, product_info_moniteur])
X, y = create_features_and_labels(df)

# Entraînement du modèle de régression
model, score, mae = train_linear_regression(X, y)

# Prédiction du prix de vente en fonction de l'estimation
estimations = np.arange(0, 2000, 50)
predictions = model.predict(create_feature_matrix(estimations))

# Génération du graphique de prédiction des prix de vente
plt.plot(X[:, 0], y, 'o')
plt.plot(estimations, predictions)
plt.title('Prédiction des prix de vente en fonction de l\'estimation')
plt.xlabel('Estimation (en euros)')
plt.ylabel('Prix de vente (en euros)')
plt.legend(['Données d\'entraînement', 'Prédictions du modèle'])
plt.show()

# Fonction pour préparer les données pour l'entraînement du modèle
def prepare_data(product_info_list):
    # Conversion des données des produits en DataFrame
    df = pd.DataFrame(product_info_list)

    # Suppression des colonnes non utilisées
    df = df.drop(['name', 'url', 'seller_name', 'end_date'], axis=1)

    # Traitement des données manquantes
    df = df.fillna(value=0)

    # Encodage des caractéristiques techniques en vecteurs de nombres
    vectorizer = CountVectorizer(token_pattern=r'[a-z]+', stop_words='english')
    X_tech = vectorizer.fit_transform(df['tech_data'])
    lda = LatentDirichletAllocation(n_components=10, random_state=42)
    X_lda = lda.fit_transform(X_tech)

    # Conversion de la colonne des prix de vente en nombres
    df['price'] = pd.to_numeric(df['price'].str.replace(r'[^\d.]', ''))

    return df, X_lda

# Fonction pour créer les caractéristiques et les étiquettes à partir des données préparées
def create_features_and_labels(df):
    # Séparation des caractéristiques et de l'étiquette
    X = df.drop(['price'], axis=1)
    y = df['price']

    return X.values, y.values

# Enregistrement des données dans un fichier Excel
file_name = 'produits.xlsx'
df.to_excel(file_name, index=False)

# Création d'un graphique dans le fichier Excel
wb = openpyxl.load_workbook(file_name)
ws = wb.active

chart = LineChart()
chart.title = 'Prédiction des prix de vente en fonction de l\'estimation'
chart.x_axis.title = 'Estimation (en euros)'
chart.y_axis.title = 'Prix de vente (en euros)'

x_data = Reference(ws, min_col=1, min_row=2, max_row=len(estimations)+1)
y_data = Reference(ws, min_col=2, min_row=1, max_row=len(estimations)+1)
series = Series(y_data, x_data, title_from_data=True)
chart.append(series)

ws.add_chart(chart, 'E2')
wb.save(file_name)

# Entraînement d'un modèle de régression avec forêt aléatoire
rf_regressor, rf_score, rf_mae = train_random_forest(X, y)

# Prédiction du prix de vente en fonction de l'estimation
estimations = np.arange(0, 2000, 50)
predictions = rf_regressor.predict(create_feature_matrix(estimations))

# Génération du graphique de prédiction des prix de vente
plt.plot(X[:, 0], y, 'o')
plt.plot(estimations, predictions)
plt.title('Prédiction des prix de vente en fonction de l\'estimation')
plt.xlabel('Estimation (en euros)')
plt.ylabel('Prix de vente (en euros)')
plt.legend(['Données d\'entraînement', 'Prédictions du modèle'])
plt.show()

# Fonction pour entraîner un modèle de régression avec forêt aléatoire
def train_random_forest(X, y, test_size=0.3, n_estimators=100, max_depth=None):
    # Séparation des données en ensembles d'entraînement et de test
    X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=test_size, random_state=42)

    # Entraînement du modèle de régression
    rf = RandomForestRegressor(n_estimators=n_estimators, max_depth=max_depth, random_state=42)
    rf.fit(X_train, y_train)

    # Calcul du coefficient de détermination et de l'erreur absolue moyenne
    rf_score = rf.score(X_test, y_test)
    rf_predictions = rf.predict(X_test)
    rf_mae = mean_absolute_error(y_test, rf_predictions)

    return rf, rf_score, rf_mae

# Écriture des données dans un fichier Excel
workbook = openpyxl.Workbook()
worksheet = workbook.active

# Écriture des en-têtes de colonnes
headers = ['Titre', 'Prix', 'État', 'Vendeur', 'Note du vendeur', 'Fin de l\'enchère']
for col_num, header in enumerate(headers, 1):
    cell = worksheet.cell(row=1, column=col_num)
    cell.value = header

# Écriture des données des produits dans les lignes suivantes
for row_num, product_info in enumerate([product_info_enchere, product_info_moniteur], 2):
    row = []
    for header in headers:
        if header == 'Titre':
            row.append(product_info['title'])
        elif header == 'Prix':
            row.append(product_info['price'])
        elif header == 'État':
            row.append(product_info['state'])
        elif header == 'Vendeur':
            row.append(product_info['seller_name'])
        elif header == 'Note du vendeur':
            row.append(product_info['seller_rating'])
        elif header == 'Fin de l\'enchère':
            row.append(product_info['end_date'])
    for col_num, value in enumerate(row, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = value

# Écriture du graphique dans le fichier Excel
chart = LineChart()
chart.title = 'Prédiction des prix de vente en fonction de l\'estimation'
chart.x_axis.title = 'Estimation (en euros)'
chart.y_axis.title = 'Prix de vente (en euros)'

x_data = Reference(worksheet, min_col=1, min_row=2, max_row=3)
y_data = Reference(worksheet, min_col=2, min_row=2, max_row=3)
chart.add_data(y_data, titles_from_data=True)
chart.set_categories(x_data)

worksheet.add_chart(chart, 'F1')

workbook.save('products.xlsx')


# Fonction pour exporter les données dans un fichier Excel
def export_to_excel(product_info_list, filename):
    # Conversion des données des produits en DataFrame
    df = pd.DataFrame(product_info_list)

    # Création d'un classeur Excel
    wb = openpyxl.Workbook()

    # Ajout des données dans une feuille Excel
    ws = wb.active
    ws.append(df.columns.tolist())
    for row in df.values:
        ws.append(list(row))

    # Ajout d'un graphique dans la feuille Excel
    chart = LineChart()
    chart.title = 'Prix de vente des produits'
    chart.x_axis.title = 'Produits'
    chart.y_axis.title = 'Prix de vente (en euros)'

    # Ajout des données dans le graphique
    data = Reference(ws, min_col=2, min_row=1, max_col=2, max_row=len(product_info_list)+1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(Reference(ws, min_col=1, min_row=2, max_row=len(product_info_list)+1))

    # Ajout du graphique dans la feuille Excel
    ws.add_chart(chart, 'D1')

    # Enregistrement du classeur Excel
    wb.save(filename)

# Fonction principale
def main():
    # Récupération des informations sur les produits
    url_enchere = 'https://www.interencheres.com/materiels-professionnels/vente-cause-cessation-magasin-dalarme-incendie-bureau-et-informatique-288053/'
    anti_captcha_key = 'XXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXX'  # Remplacer par une clé AntiCaptcha valide
    product_info_enchere = get_product_info_interencheres(url_enchere, anti_captcha_key)

    url_moniteur = 'https://www.moniteurlive.com/vente/52456/lot/11671073'
    product_info_moniteur = get_product_info_moniteurlive(url_moniteur)

    # Préparation des données pour l'entraînement du modèle
    df = prepare_data([product_info_enchere, product_info_moniteur])
    X, y = create_features_and_labels(df)

    # Entraînement du modèle de régression
    model, score, mae = train_linear_regression(X, y)

    # Prédiction du prix de vente en fonction de l'estimation
    estimations = np.arange(0, 2000, 50)
    predictions = model.predict(create_feature_matrix(estimations))

    # Affichage du score et de l'erreur moyenne absolue du modèle
    print(f'Score du modèle : {score:.2f}')
    print(f'Erreur moyenne absolue : {mae:.2f}')

    # Affichage des prédictions du modèle
    for estimation, prediction in zip(estimations, predictions):
        print(f'Estimation : {estimation} € / Prédiction : {prediction:.2f} €')

    # Génération du graphique de prédiction des prix de vente
    plt.plot(X[:, 0], y, 'o')
    plt.plot(estimations, predictions)
    plt.title('Prédiction des prix de vente en fonction de l\'estimation')
    plt.xlabel('Estimation (en euros)')
    plt.ylabel('Prix de vente (en euros)')
    plt.legend(['Données d\'entraînement', 'Prédictions du model'])
    plt.show()

if __name__ == '__main__':
    main()